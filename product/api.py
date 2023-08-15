import csv
import json
import datetime
# import requests
# import uuid
# import time
import openpyxl

from django.conf import settings
# from io import BytesIO
# from PIL import Image
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet
# from rest_framework.filters import SearchFilter
from dry_rest_permissions.generics import DRYPermissions

from ebaysdk.trading import Connection
from .models import Product, DeletedList, OrderList
from .serializers import ProductSerializer
from product.scrape.engineselector import select_engine
from utils.convertcurrency import getCurrentRate
# from utils.ebay_policy import DISPATCHTIMEMAX, RETURN_POLICY, SHIPPING_POLICY
# from .filterbackend import FilterBackend

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from utils.profit_formula import profit_formula

class ProductViewSet(ModelViewSet):
    serializer_class = ProductSerializer
    queryset = Product.objects.all()
    permission_classes = (DRYPermissions, )

    @action(detail=False, methods=['POST'])
    def scrape_data(self, request):
        url = request.data['url']
        engine = select_engine(url=url)
        if engine:
            engine = engine()
            try:
                data = engine.scrape_data(source_url=url)
                # data['price_en'] = convert('JPY', 'USD', data['price_jp'])

                return Response(
                    data=data,
                    status=200
                )
            except Exception as err:
                raise err
        
        return Response(
            data='入力したサイトへのサービスはまだサポートされていません。',
            status=400
        )
            
    @action(detail=False, methods=['POST'])
    def validate_product(self, request):
        info = request.data['product_info']
        purchase_url = info['purchase_url']
        ebay_url = info['ebay_url']
        mode = info['mode']

        if mode == 1:
            try:
                # validate duplicate
                products = Product.objects.filter(purchase_url = purchase_url).values() | Product.objects.filter(ebay_url = ebay_url).values()

            except Exception as err:
                return Response(
                    data = 'すでに存在しています！',
                    status = 401
                )


            if len(products) > 0:
                return Response(
                    data = 'すでに存在しています！',
                    status = 401
                )

        engine = select_engine(url = purchase_url)
        
        if engine:
            engine = engine()
            try:
                data = engine.scrape_data(source_url = purchase_url)

                if data['nothing'] == False:
                    # data['sell_price_en'] = convert('JPY', 'USD', data['purchase_price'])

                    return Response(
                        data = data,
                        status = 200
                    )
                
                else:
                    return Response(
                        data = 'この商品は削除されました。',
                        status = 401
                    )
            except Exception as err:
                raise err
        else:
            return Response(
                data = '入力したサイトへのサービスはまだサポートされていません。',
                status = 401
            )
        
    @action(detail=False, methods=['GET'])
    def get_products(self, request):

        # perPageNum = request.GET.get('pageSize')
        # page = request.GET.get('page')
        creator = request.GET.get('created_by')
        superuser = request.GET.get('is_superuser')

        products = []

        if creator == '':
            if superuser == 'true':
                products = Product.objects.all().order_by('-id')
        else:
            products = Product.objects.filter(created_by = creator).order_by('-id')

        return Response(
            data = products.values(),
            status = 200
        )
    
    @action(detail=False, methods=['POST'])
    def add_item(self, request):
        item = request.data['product']
        mode = request.data['mode']
        product = ()

        if mode == 1:
            try:
                product = Product(
                    created_at = item['created_at'],
                    product_name = item['product_name'],
                    ec_site = item['ec_site'],
                    purchase_url = item['purchase_url'],
                    ebay_url = item['ebay_url'],
                    purchase_price = item['purchase_price'],
                    sell_price_en = item['sell_price_en'],
                    profit = item['profit'],
                    profit_rate = item['profit_rate'],
                    prima = item['prima'],
                    shipping = item['shipping'],
                    quantity = item['quantity'],
                    created_by = item['created_by'],
                    notes = item['notes']
                )
                
                product.save()

                return Response(
                    {'Success!'},
                    status = 200
                )
                
            except:
                return Response(
                    data = '登録操作が失敗しました',
                    status = 401
                )
        else:
            try:
                pid = item['id']
                user = Product.objects.filter(id = pid)
                user.update(
                    created_at = item['created_at'],
                    product_name = item['product_name'],
                    ec_site = item['ec_site'],
                    purchase_url = item['purchase_url'],
                    ebay_url = item['ebay_url'],
                    purchase_price = item['purchase_price'],
                    sell_price_en = item['sell_price_en'],
                    profit = item['profit'],
                    profit_rate = item['profit_rate'],
                    prima = item['prima'],
                    shipping = item['shipping'],
                    quantity = item['quantity'],
                    created_by = item['created_by'],
                    notes = item['notes']
                )

                return Response(
                    {'Success!'},
                    status = 200
                )
                
            except:
                return Response(
                    data = '編集操作が失敗しました！',
                    status = 401
                )
        
    @action(detail=False, methods=['GET'])  
    def get_results(self, request):

        worker = request.GET.get('worker')
        month = request.GET.get('month')
        results = Product.objects.filter(created_by = worker, created_at__startswith = month).order_by('-id').values('id', 'created_at', 'created_by')

        return Response({'results': results}, status = 200)

    @action(detail=False, methods=['POST'])
    def upload_product_file(self, request):
        file = request.FILES['csvFile']
        user = request.data['user']

        date = str(datetime.datetime.now()).split(" ")[0]

        wb = openpyxl.load_workbook(file)
        wproductsInfo = wb['Book1']

        for i in range(1, wproductsInfo.max_row):
            purchase_url = wproductsInfo.cell(row=i+1, column=4).value

            if len(Product.objects.filter(purchase_url = purchase_url)) > 0:
                continue

            purchase_price = str(wproductsInfo.cell(row=i+1,column=6).value).replace("¥", "").replace(",", "")
            sell_price_en = str(wproductsInfo.cell(row=i+1,column=7).value).replace("$", "").replace(",", "")
            profit = str(wproductsInfo.cell(row=i+1,column=8).value).replace("¥", "").replace(",", "")
            profit_rate = str(wproductsInfo.cell(row=i+1,column=9).value).replace(",", "").replace("%", "")
            prima = str(wproductsInfo.cell(row=i+1,column=11).value).replace("¥", "").replace(",", "")
            shipping = str(wproductsInfo.cell(row=i+1,column=12).value).replace("$", "").replace(",", "")

            if purchase_price.isnumeric() == False:
                purchase_price = 0

            if sell_price_en.isnumeric() == False:
                sell_price_en = 0

            if profit.isnumeric() == False:
                profit = 0

            if profit_rate.isnumeric() == False:
                profit_rate = 0

            if prima.isnumeric() == False:
                prima = 0

            if shipping.isnumeric() == False:
                shipping = 0


            product = Product(
                created_at = date,
                product_name = wproductsInfo.cell(row=i+1, column=2).value,
                ec_site = wproductsInfo.cell(row=i+1, column=3).value,
                purchase_url = purchase_url,
                ebay_url = wproductsInfo.cell(row=i+1, column=5).value,
                purchase_price = purchase_price,
                sell_price_en = sell_price_en,
                profit = profit,
                profit_rate = profit_rate,
                prima = prima,
                shipping = shipping,
                quantity = 0,
                created_by = user,
                notes = wproductsInfo.cell(row=i+1, column=12).value
            )

            product.save()

        return Response(
            data="success",
            status=200
        )

    @action(detail=False, methods=['POST'])
    def upload_order_file(self, request):
        file = request.FILES['csvFile']
        user = request.data['user']

        date = str(datetime.datetime.now()).split(" ")[0]

        wb = openpyxl.load_workbook(file)
        wproductsInfo = wb['Book1']

        for i in range(1, wproductsInfo.max_row):
            purchase_url = wproductsInfo.cell(row=i+1, column=4).value

            if len(OrderList.objects.filter(purchase_url = purchase_url)) > 0:
                continue

            purchase_price = str(wproductsInfo.cell(row=i+1,column=6).value).replace("¥", "").replace(",", "")
            sell_price_en = str(wproductsInfo.cell(row=i+1,column=7).value).replace("$", "").replace(",", "")
            profit = str(wproductsInfo.cell(row=i+1,column=8).value).replace("¥", "").replace(",", "")
            profit_rate = str(wproductsInfo.cell(row=i+1,column=9).value).replace(",", "").replace("%", "")
            prima = str(wproductsInfo.cell(row=i+1,column=11).value).replace("¥", "").replace(",", "")
            shipping = str(wproductsInfo.cell(row=i+1,column=12).value).replace("$", "").replace(",", "")

            if purchase_price.isnumeric() == False:
                purchase_price = 0

            if sell_price_en.isnumeric() == False:
                sell_price_en = 0

            if profit.isnumeric() == False:
                profit = 0

            if profit_rate.isnumeric() == False:
                profit_rate = 0

            if prima.isnumeric() == False:
                prima = 0

            if shipping.isnumeric() == False:
                shipping = 0


            product = OrderList(
                created_at = date,
                product_name = wproductsInfo.cell(row=i+1, column=2).value,
                ec_site = wproductsInfo.cell(row=i+1, column=3).value,
                purchase_url = purchase_url,
                ebay_url = wproductsInfo.cell(row=i+1, column=5).value,
                purchase_price = purchase_price,
                sell_price_en = sell_price_en,
                profit = profit,
                profit_rate = profit_rate,
                prima = prima,
                shipping = shipping,
                quantity = 0,
                created_by = user,
                notes = wproductsInfo.cell(row=i+1, column=12).value
            )

            product.save()

        return Response(
            data="success",
            status=200
        )
    
        
    @action(detail=False, methods=['GET'])    
    def get_orders(self, request):
        orders = OrderList.objects.all().order_by('-id')

        return Response(
            data = orders.values(),
            status = 200
        )
    
    @action(detail=False, methods=['POST'])
    def add_order_item(self, request):
        item = request.data['order']
        mode = request.data['mode']

        if mode == 1:
            try:
                order = OrderList(
                    created_at = item['created_at'],
                    product_name = item['product_name'],
                    ec_site = item['ec_site'],
                    purchase_url = item['purchase_url'],
                    ebay_url = item['ebay_url'],
                    purchase_price = item['purchase_price'],
                    sell_price_en = item['sell_price_en'],
                    profit = item['profit'],
                    profit_rate = item['profit_rate'],
                    prima = item['prima'],
                    shipping = item['shipping'],
                    quantity = item['quantity'],
                    order_num = item['order_num'],
                    ordered_at = item['ordered_at'],
                    created_by = item['created_by'],
                    notes = item['notes']
                )
                
                order.save()

                return Response(
                    {'Success!'},
                    status=200
                )
            
            except:
                return Response(
                    data = 'オーダー商品登録作業が失敗しました！',
                    status = 401
                )
        else:
            try:
                pid = item['id']
                order = OrderList.objects.filter(id = pid)

                order.update(
                    created_at = item['created_at'],
                    product_name = item['product_name'],
                    ec_site = item['ec_site'],
                    purchase_url = item['purchase_url'],
                    ebay_url = item['ebay_url'],
                    purchase_price = item['purchase_price'],
                    sell_price_en = item['sell_price_en'],
                    profit = item['profit'],
                    profit_rate = item['profit_rate'],
                    prima = item['prima'],
                    shipping = item['shipping'],
                    quantity = item['quantity'],
                    order_num = item['order_num'],
                    ordered_at = item['ordered_at'],
                    created_by = item['created_by'],
                    notes = item['notes']
                )

                return Response(
                    {'Success!'},
                    status=200
                )
            
            except:
                return Response(
                    data = 'オーダー商品登録作業が失敗しました！',
                    status = 401
                )

    @action(detail=False, methods=['GET'])  
    def get_deleted_items(self, request):

        # perPageNum = request.GET.get('pageSize')
        # page = request.GET.get('page')

        deletes_list = DeletedList.objects.all().order_by('-id')

        # paginator = Paginator(deletes_list, perPageNum)

        # try:
        #     deletes = paginator.page(page).object_list
        # except PageNotAnInteger:
        #     deletes = paginator.page(1).object_list
        # except EmptyPage:
        #     deletes = paginator.page(paginator.num_pages).object_list

        # return Response({'count': paginator.count, 'deleted_items': deletes.values()}, status=200)

        return Response(
            data = deletes_list.values(),
            status = 200
        )
    
    @action(detail=False, methods=['POST'])  
    def delete_product(self, request):
        id = request.data['id']
        
        item = Product.objects.filter(id = id)[0]

        try:
            date = datetime.datetime.now()
            del_item = ()

            del_item = DeletedList(
                    created_at = item.created_at,
                    updated_at = "",
                    product_name = item.product_name,
                    ec_site = item.ec_site,
                    purchase_url = item.purchase_url,
                    ebay_url = item.ebay_url,
                    purchase_price = item.purchase_price,
                    sell_price_en = item.sell_price_en,
                    profit = item.profit,
                    profit_rate = item.profit_rate,
                    prima = item.prima,
                    shipping = item.shipping,
                    quantity = item.quantity,
                    created_by = item.created_by,
                    notes = item.notes,
                    deleted_at = date
                )
            
            del_item.save()
            item.delete()

            return Response(data='success!', status=200)
        except:
            return Response(
                data = '削除操作が失敗しました！',
                status = 401
            )

    @action(detail=False, methods=['POST'])
    def delete_order_item(self, request):
        pid = request.data['id']

        try:
            OrderList.objects.get(id = pid).delete()

            return Response(data='success!', status=200)
        except:
            return Response(
                data = '削除操作が失敗しました！',
                status = 401
            )
    
    @action(detail=False, methods=['GET'])
    def shipping_fee(self, request):
        with open(file=str(settings.BASE_DIR / 'utils/shipping_fee.txt'),  mode='r', encoding='utf-8') as f:
            fee = f.read()
        
        return Response(
            data=json.loads(fee),
            status=200
        )
    
    @action(detail=False, methods=['GET'])
    def settings_attr(self, request):
        with open(file=str(settings.BASE_DIR / 'utils/settings_attrs.txt'),  mode='r', encoding='utf-8') as f:
            settings_attrs = f.read()

        res = json.loads(settings_attrs)
        rate = getCurrentRate('JPY')

        
        res['rate'] = str(rate)

        with open(file=str(settings.BASE_DIR / 'utils/settings_attrs.txt'),  mode='w', encoding='utf-8') as f:
            f.write(json.dumps(res, indent=4))

        # self.updateProductInfo()
        
        return Response(
            data = res,
            status = 200
        )
    
    @action(detail=False, methods=['POST'])
    def update_settings_attr(self, request):
        settings_attr = request.data['settings_attr']

        # st = json.dumps(settings_attr, indent=4)

        with open(file=str(settings.BASE_DIR / 'utils/settings_attrs.txt'),  mode='w', encoding='utf-8') as f:
            f.write(json.dumps(settings_attr, ensure_ascii=False, indent=4))

        return Response(
            data=settings_attr,
            status=200
        )
    
    @action(detail=False, methods=['GET'])
    def get_ecsites(self, request):
        with open(file=str(settings.BASE_DIR / 'utils/scrape_site.txt'),  mode='r', encoding='utf-8') as f:
            ecsites = f.read()
        
        return Response(
            data=json.loads(ecsites),
            status=200
        )
    
    @action(detail=False, methods=['POST'])
    def update_ecsites(self, request):
        ecsites = request.data['ecsites']
        with open(file=str(settings.BASE_DIR / 'utils/scrape_site.txt'),  mode='w', encoding='utf-8') as f:
            f.write(json.dumps(ecsites, indent=4))

        return Response(
            data='Success',
            status=200
        )

    
    @action(detail=False, methods=['POST'])
    def download_product(self, request):
        with open(file=str(settings.BASE_DIR / 'media/products.csv'),  mode='w', encoding='utf-8', newline='') as f:
            fieldnames = ['商品名', 'EC site', '仕入れ URL', 'eBay URL', '利益額', '利益率', '仕入価格（円）', 'フリマ送料', '仕入合計（円）', '販売価格', '輸出送料', '出品者', '備考']
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            
            for product in Product.objects.all():
                writer.writerow({
                    '商品名': product.product_name,
                    'EC site': product.ec_site, 
                    '仕入れ URL': product.purchase_url, 
                    'eBay URL': product.ebay_url, 
                    '利益額': product.profit, 
                    '利益率': product.profit_rate, 
                    '仕入価格（円）': product.purchase_price, 
                    'フリマ送料': product.prima, 
                    '仕入合計（円）': product.purchase_price + product.prima, 
                    '販売価格': product.sell_price_en, 
                    '輸出送料': product.shipping, 
                    '出品者': product.created_by, 
                    '備考': product.notes
                })
        return Response(
            data='Success',
            status=200
        )
    
    @action(detail=False, methods=['GET'])
    def update_info(self, request):
        with open(file=str(settings.BASE_DIR / 'utils/settings_attrs.txt'),  mode='r', encoding='utf-8') as f:
            settings_attrs = f.read()

        res = json.loads(settings_attrs)
        rate = getCurrentRate('JPY')
        res['rate'] = str(rate)

        with open(file=str(settings.BASE_DIR / 'utils/settings_attrs.txt'),  mode='w', encoding='utf-8') as f:
            f.write(json.dumps(res, indent=4))

        # update profits
        products = Product.objects.all().order_by('-id')

        for product in products:
            sell_price = float(product.sell_price_en)
            purchase_price = product.purchase_price
            prima = product.prima
            shipping = product.shipping

            profit = profit_formula(float(sell_price), int(purchase_price), float(prima), float(shipping), res)
            profit_rate = (profit / (sell_price * rate)) * 100

            Product.objects.filter(id = product.id).update(profit = profit, profit_rate = profit_rate)
        
        return Response(
            data = res,
            status = 200
        )

    @action(detail=False, methods=['POST'])
    def get_item_specific(self, request):
        item_number = request.data['item_number']

        try:
            # Set up the API connection
            api = Connection(appid=request.user.app_id, devid=request.user.dev_id, certid=request.user.cert_id, token=request.user.ebay_token, config_file=None)
            response = api.execute(
                'GetItem',
                {
                    'ItemID': item_number,
                    'IncludeItemSpecifics': 'True'
                }
            )
            item_specifics = response.reply.Item.ItemSpecifics.NameValueList
            required_specifics = []
            optional_specifics = []
            for item_specific in item_specifics:
                if(getattr(item_specific, 'Name') == 'Brand' or getattr(item_specific, 'Name') == 'Type'):
                    required_specifics.append(
                        {
                            'Name': item_specific.Name,
                            'Value': item_specific.Value,
                            'Condition': 'Required',
                        }
                    )
                else:
                    optional_specifics.append(
                        {
                            'Name': item_specific.Name,
                            'Value': item_specific.Value,
                            'Condition': 'Optional',
                        }
                    )
            result = required_specifics + optional_specifics
            return Response(
                data = result,
                status = 200
            )
        except Exception as err:
            raise err

            
